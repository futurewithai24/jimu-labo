"""
kintai_mail_collector.py
========================
タスクスケジューラから毎朝6:00に自動実行される。
Outlookの受信トレイから当日のメールを取得し、
出勤記録.xlsx の日付シートに自動記録する。

【セットアップ手順】
1. ライブラリインストール
   pip install pywin32 openpyxl

2. OUTPUT_EXCEL_PATH を実際のパスに変更する

3. タスクスケジューラに登録する（後述）

【タスクスケジューラ登録手順】
1. スタートメニュー →「タスクスケジューラ」を開く
2. 右側「基本タスクの作成」をクリック
3. 名前：「出勤記録自動集計」→ 次へ
4. トリガー：「毎日」→ 次へ
5. 開始時刻：6:00:00 → 次へ
6. 操作：「プログラムの開始」→ 次へ
7. プログラム：python.exe のフルパス
   例: C:\\Users\\ユーザー名\\AppData\\Local\\Programs\\Python\\Python311\\python.exe
8. 引数：このファイルのフルパス
   例: C:\\Users\\ユーザー名\\Documents\\kintai_app\\kintai_mail_collector.py
9. 完了
"""

import win32com.client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import sys

# ============================================================
# ★ 設定：ここだけ変更してください
# ============================================================

# 出勤記録Excelの保存先（デスクトップ）
OUTPUT_EXCEL_PATH = os.path.join(os.path.expanduser("~"), "Desktop", "出勤記録.xlsx")

# ============================================================
# メイン処理
# ============================================================

def get_target_date():
    """処理対象日（当日）を返す"""
    return datetime.now().date()


def get_sender_name(mail):
    """送信者の表示名を取得。取得できなければメールアドレスを返す"""
    try:
        name = mail.SenderName
        if name:
            return name.strip()
    except Exception:
        pass
    try:
        return mail.SenderEmailAddress.strip()
    except Exception:
        return "不明"


def fetch_mails_from_outlook(target_date):
    """
    Outlookの受信トレイから当日のメールを取得し
    [(送信者名, 送信時刻)] のリストを返す
    """
    print(f"Outlookに接続中...")
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    inbox = namespace.GetDefaultFolder(6)  # 6 = 受信トレイ

    print(f"対象日: {target_date}　メール取得中...")
    results = []

    for mail in inbox.Items:
        try:
            # olMailItem (Class=43) のみ処理。会議通知・タスク等はスキップ
            if not hasattr(mail, 'Class') or mail.Class != 43:
                continue

            received = mail.ReceivedTime
            mail_date = datetime(
                received.year,
                received.month,
                received.day
            ).date()

            if mail_date != target_date:
                continue

            # 件名に「勤怠連絡」を含むメールのみ対象
            subject = mail.Subject or ""
            if "勤怠連絡" not in subject:
                continue

            sender_name = get_sender_name(mail)
            mail_time = f"{received.hour:02d}:{received.minute:02d}"
            results.append((sender_name, mail_time))
            print(f"  取得: {sender_name} / {mail_time}")

        except Exception as e:
            print(f"  スキップ（エラー）: {e}")
            continue

    print(f"取得完了: {len(results)}件")
    return results


def get_or_create_workbook(path):
    """Excelファイルがあれば開く、なければ新規作成"""
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        print(f"新規Excelファイルを作成: {path}")
        return wb


def get_or_create_sheet(wb, sheet_name):
    """シートがあれば返す、なければ新規作成してヘッダーを設定"""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(["氏名", "出勤時刻"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.row_dimensions[1].height = 20

    print(f"新規シート作成: {sheet_name}")
    return ws


def write_to_excel(path, target_date, mail_records):
    """
    対象日のシートにメールデータを書き込む
    同一送信者が既に記録されている場合はスキップ（重複防止）
    """
    wb = get_or_create_workbook(path)
    sheet_name = f"{target_date.month}/{target_date.day}"
    ws = get_or_create_sheet(wb, sheet_name)

    # 既存データの送信者名を取得（重複チェック用）
    existing_names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_names.add(row[0])

    data_font = Font(name="Meiryo UI", size=11)
    center_align = Alignment(horizontal="center")
    added_count = 0
    skip_count = 0

    for sender_name, mail_time in mail_records:
        if sender_name in existing_names:
            print(f"  スキップ（重複）: {sender_name}")
            skip_count += 1
            continue

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=sender_name).font = data_font
        time_cell = ws.cell(row=next_row, column=2, value=mail_time)
        time_cell.font = data_font
        time_cell.alignment = center_align
        existing_names.add(sender_name)
        added_count += 1
        print(f"  書き込み: {sender_name} / {mail_time}")

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"\n保存完了: {path}")
    print(f"追加: {added_count}件　スキップ（重複）: {skip_count}件")


def main():
    print("=" * 50)
    print(f"出勤メール自動集計　開始: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}")
    print("=" * 50)

    target_date = get_target_date()
    mail_records = fetch_mails_from_outlook(target_date)

    if not mail_records:
        print(f"\n{target_date} の受信メールは0件でした。処理を終了します。")
        sys.exit(0)

    write_to_excel(OUTPUT_EXCEL_PATH, target_date, mail_records)
    print(f"\n完了: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
