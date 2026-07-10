# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

st.set_page_config(page_title="経費集計ツール", page_icon="💴", layout="wide")
st.title("💴 経費集計ツール")
st.caption("経費を入力するだけで、カテゴリ別に自動集計 → Excelでダウンロードできます。無料・登録不要。")

CATEGORIES = ["交通費", "食費", "宿泊費", "消耗品", "通信費", "接待費", "その他"]
PAYMENTS   = ["現金", "クレジットカード", "電子マネー"]
COLS       = ["日付", "カテゴリ", "項目名", "金額", "支払方法", "備考"]

if "expenses" not in st.session_state:
    st.session_state.expenses = pd.DataFrame(columns=COLS)

# ─────────────────────────────────────────────
# ① 経費を入力
# ─────────────────────────────────────────────
st.subheader("① 経費を入力")

with st.form("add_expense", clear_on_submit=True):
    c1, c2, c3 = st.columns(3)
    with c1:
        f_date     = st.date_input("日付", value=date.today())
        f_category = st.selectbox("カテゴリ", CATEGORIES)
    with c2:
        f_item    = st.text_input("項目名", placeholder="例：電車代（新宿→品川）")
        f_payment = st.selectbox("支払方法", PAYMENTS)
    with c3:
        f_amount = st.number_input("金額（円）", min_value=0, step=100)
        f_note   = st.text_input("備考", placeholder="任意")

    if st.form_submit_button("➕ 追加する", use_container_width=True, type="primary"):
        if not f_item:
            st.warning("項目名を入力してください。")
        elif f_amount <= 0:
            st.warning("金額を入力してください。")
        else:
            new_row = pd.DataFrame([{
                "日付":    f_date.strftime("%Y/%m/%d"),
                "カテゴリ": f_category,
                "項目名":  f_item,
                "金額":    int(f_amount),
                "支払方法": f_payment,
                "備考":    f_note,
            }])
            st.session_state.expenses = pd.concat(
                [st.session_state.expenses, new_row], ignore_index=True
            )
            st.success(f"追加しました：{f_item}（¥{int(f_amount):,}）")

df = st.session_state.expenses.copy()
if df.empty:
    st.info("経費を入力すると、明細と集計がここに表示されます。")
    st.stop()

st.divider()

# ─────────────────────────────────────────────
# ② 明細一覧（編集・削除可能）
# ─────────────────────────────────────────────
st.subheader("② 明細一覧")

months    = ["すべて"] + sorted(df["日付"].str[:7].unique().tolist(), reverse=True)
sel_month = st.selectbox("月を選択", months)
is_all    = sel_month == "すべて"
df_view   = df.copy() if is_all else df[df["日付"].str.startswith(sel_month)].copy()

edited = st.data_editor(
    df_view.reset_index(drop=True),
    use_container_width=True,
    num_rows="dynamic",
    column_config={
        "金額":    st.column_config.NumberColumn("金額（円）", format="%d円", min_value=0),
        "カテゴリ": st.column_config.SelectboxColumn("カテゴリ", options=CATEGORIES, required=True),
        "支払方法": st.column_config.SelectboxColumn("支払方法", options=PAYMENTS, required=True),
    },
    hide_index=True,
)

if is_all:
    st.session_state.expenses = edited.copy()
else:
    other = df[~df["日付"].str.startswith(sel_month)].copy()
    st.session_state.expenses = pd.concat([other, edited], ignore_index=True)

df      = st.session_state.expenses.copy()
df_view = df.copy() if is_all else df[df["日付"].str.startswith(sel_month)].copy()

if st.button("🗑️ すべてクリア", type="secondary"):
    st.session_state.expenses = pd.DataFrame(columns=COLS)
    st.rerun()

st.divider()

# ─────────────────────────────────────────────
# ③ カテゴリ別集計
# ─────────────────────────────────────────────
st.subheader("③ カテゴリ別集計")

df_view["金額"] = pd.to_numeric(df_view["金額"], errors="coerce").fillna(0)
summary = (
    df_view.groupby("カテゴリ", sort=False)["金額"]
    .sum()
    .reset_index()
    .rename(columns={"金額": "合計金額"})
)
summary = summary[summary["合計金額"] > 0].sort_values("合計金額", ascending=False)
total   = int(df_view["金額"].sum())

cl, cr = st.columns([1, 1])

with cl:
    disp = summary.copy()
    disp["合計金額"] = disp["合計金額"].apply(lambda x: f"¥{int(x):,}")
    disp.columns = ["カテゴリ", "合計金額（円）"]
    st.dataframe(disp, use_container_width=True, hide_index=True)
    st.metric("合計金額", f"¥{total:,}")

with cr:
    if not summary.empty:
        st.bar_chart(summary.set_index("カテゴリ")["合計金額"])

st.divider()

# ─────────────────────────────────────────────
# ④ Excelダウンロード
# ─────────────────────────────────────────────
st.subheader("④ Excelダウンロード")

period_label = sel_month if not is_all else "全期間"
st.caption(f"対象：{period_label}　明細 {len(df_view)} 件")


def build_excel(detail_df, summary_df, total_amount, label):
    wb = openpyxl.Workbook()

    h_fill  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    h_font  = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=11)
    t_fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    t_font  = Font(name="Meiryo UI", bold=True, size=11)
    n_font  = Font(name="Meiryo UI", size=11)
    c_align = Alignment(horizontal="center", vertical="center")
    r_align = Alignment(horizontal="right",  vertical="center")

    ws1       = wb.active
    ws1.title = "カテゴリ別集計"
    ws1["A1"] = f"経費集計　{label}"
    ws1["A1"].font = Font(name="Meiryo UI", bold=True, size=13)
    ws1.row_dimensions[1].height = 24

    for col, h in enumerate(["カテゴリ", "合計金額（円）"], 1):
        c = ws1.cell(row=3, column=col, value=h)
        c.fill, c.font, c.alignment = h_fill, h_font, c_align
    ws1.row_dimensions[3].height = 20

    for r, row in enumerate(summary_df.itertuples(index=False), 4):
        ws1.cell(row=r, column=1, value=row[0]).font = n_font
        amt = ws1.cell(row=r, column=2, value=int(row[1]))
        amt.font, amt.number_format, amt.alignment = n_font, "#,##0", r_align

    tr = 4 + len(summary_df)
    ws1.cell(row=tr, column=1, value="合計").font = t_font
    ws1.cell(row=tr, column=1).fill = t_fill
    tc = ws1.cell(row=tr, column=2, value=total_amount)
    tc.font, tc.fill, tc.number_format, tc.alignment = t_font, t_fill, "#,##0", r_align

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 18

    ws2       = wb.create_sheet(title="明細一覧")
    d_headers = ["日付", "カテゴリ", "項目名", "金額（円）", "支払方法", "備考"]
    widths    = [14, 14, 30, 16, 18, 20]

    for col, (h, w) in enumerate(zip(d_headers, widths), 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = h_fill, h_font, c_align
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 20

    for r, row in enumerate(detail_df.itertuples(index=False), 2):
        values = [
            str(row[0]), str(row[1]), str(row[2]),
            int(row[3]) if str(row[3]).replace(".", "").isdigit() else 0,
            str(row[4]), str(row[5]) if row[5] else "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.font = n_font
            if col == 4:
                cell.number_format = "#,##0"
                cell.alignment = r_align

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


excel_buf = build_excel(df_view, summary, total, period_label)
file_name = f"経費集計_{period_label.replace('/', '')}.xlsx"

st.download_button(
    label="📥 Excelをダウンロード",
    data=excel_buf,
    file_name=file_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    type="primary",
)
