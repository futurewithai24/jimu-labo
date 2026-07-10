# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="出張経費精算書", page_icon="✈️", layout="wide")
st.title("✈️ 出張経費精算書")
st.caption("出張情報と経費を入力するだけで、精算書Excelを自動作成します。登録不要・無料。")

CATEGORIES = ["交通費", "食費", "宿泊費", "消耗品", "通信費", "接待費", "その他"]
PAYMENTS   = ["現金", "クレジットカード", "電子マネー"]
COLS       = ["日付", "カテゴリ", "項目名", "金額", "支払方法", "備考"]

if "expenses" not in st.session_state:
    st.session_state.expenses = pd.DataFrame(columns=COLS)

# ─────────────────────────────────────────────
# ① 出張情報の入力
# ─────────────────────────────────────────────
st.subheader("① 出張情報")

c1, c2 = st.columns(2)
with c1:
    applicant   = st.text_input("申請者名", placeholder="例：まるのまい")
    destination = st.text_input("出張先", placeholder="例：新神戸")
with c2:
    purpose  = st.text_input("出張目的", placeholder="例：展示会・営業訪問")
    dc1, dc2 = st.columns(2)
    with dc1:
        trip_from = st.date_input("出張期間（開始）", value=date.today())
    with dc2:
        trip_to   = st.date_input("出張期間（終了）",  value=date.today())

st.divider()

# ─────────────────────────────────────────────
# ② 経費を入力
# ─────────────────────────────────────────────
st.subheader("② 経費を入力")

with st.form("add_expense", clear_on_submit=True):
    c1, c2, c3 = st.columns(3)
    with c1:
        f_date     = st.date_input("日付", value=date.today())
        f_category = st.selectbox("カテゴリ", CATEGORIES)
    with c2:
        f_item    = st.text_input("項目名", placeholder="例：電車代（東京→新神戸）")
        f_payment = st.selectbox("支払方法", PAYMENTS)
    with c3:
        f_amount = st.number_input("金額（円）", min_value=0, step=100)
        f_note   = st.text_input("備考", placeholder="任意")

    if st.form_submit_button("➕ 追加する", use_container_width=True, type="primary"):
        if not f_item:
            st.warning("項目名を入力してください。")
        elif f_amount <= 0:
            st.warning("金額を入力してください。")
        else:
            new_row = pd.DataFrame([{
                "日付":    f_date.strftime("%Y/%m/%d"),
                "カテゴリ": f_category,
                "項目名":  f_item,
                "金額":    int(f_amount),
                "支払方法": f_payment,
                "備考":    f_note,
            }])
            st.session_state.expenses = pd.concat(
                [st.session_state.expenses, new_row], ignore_index=True
            )
            st.success(f"追加しました：{f_item}（¥{int(f_amount):,}）")

df = st.session_state.expenses.copy()
if df.empty:
    st.info("経費を入力すると、明細と集計がここに表示されます。")
    st.stop()

st.divider()

# ─────────────────────────────────────────────
# ③ 明細一覧（編集・削除可能）
# ─────────────────────────────────────────────
st.subheader("③ 明細一覧")

edited = st.data_editor(
    df.reset_index(drop=True),
    use_container_width=True,
    num_rows="dynamic",
    column_config={
        "金額":    st.column_config.NumberColumn("金額（円）", format="%d円", min_value=0),
        "カテゴリ": st.column_config.SelectboxColumn("カテゴリ", options=CATEGORIES, required=True),
        "支払方法": st.column_config.SelectboxColumn("支払方法", options=PAYMENTS, required=True),
    },
    hide_index=True,
)
st.session_state.expenses = edited.copy()
df = st.session_state.expenses.copy()

if st.button("🗑️ すべてクリア", type="secondary"):
    st.session_state.expenses = pd.DataFrame(columns=COLS)
    st.rerun()

st.divider()

# ─────────────────────────────────────────────
# ④ カテゴリ別集計
# ─────────────────────────────────────────────
st.subheader("④ カテゴリ別集計")

df["金額"] = pd.to_numeric(df["金額"], errors="coerce").fillna(0)
summary = (
    df.groupby("カテゴリ", sort=False)["金額"]
    .sum()
    .reset_index()
    .rename(columns={"金額": "合計金額"})
)
summary = summary[summary["合計金額"] > 0].sort_values("合計金額", ascending=False)
total   = int(df["金額"].sum())

cl, cr = st.columns([1, 1])
with cl:
    disp = summary.copy()
    disp["合計金額"] = disp["合計金額"].apply(lambda x: f"¥{int(x):,}")
    disp.columns = ["カテゴリ", "合計金額（円）"]
    st.dataframe(disp, use_container_width=True, hide_index=True)
    st.metric("合計金額", f"¥{total:,}")
with cr:
    if not summary.empty:
        st.bar_chart(summary.set_index("カテゴリ")["合計金額"])

st.divider()

# ─────────────────────────────────────────────
# ⑤ 精算書Excelをダウンロード
# ─────────────────────────────────────────────
st.subheader("⑤ 精算書Excelをダウンロード")
st.caption(f"明細 {len(df)} 件　合計 ¥{total:,}")


def build_seisan_excel(detail_df, summary_df, total_amount,
                       applicant, destination, purpose, trip_from, trip_to):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "出張経費精算書"

    thin             = Side(style="thin")
    border           = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_no_bottom  = Border(left=thin, right=thin, top=thin)
    border_tot_label  = Border(right=thin, top=thin)
    h_fill    = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sub_fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    h_font    = Font(name="Meiryo UI", bold=True, color="FFFFFF", size=11)
    sub_font  = Font(name="Meiryo UI", bold=True, size=11)
    n_font    = Font(name="Meiryo UI", size=11)
    title_font = Font(name="Meiryo UI", bold=True, size=15)
    c_align   = Alignment(horizontal="center", vertical="center")
    l_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    r_align   = Alignment(horizontal="right",  vertical="center")

    # 列幅（ユーザー指定フォーマットに合わせる）
    for col, w in zip("ABCDEFG", [15.0, 15.2, 14.0, 28.0, 14.6, 14.6, 14.6]):
        ws.column_dimensions[col].width = w

    # ── タイトル ──
    ws.merge_cells("A1:G1")
    ws["A1"] = "出張経費精算書"
    ws["A1"].font      = title_font
    ws["A1"].alignment = c_align
    ws.row_dimensions[1].height = 32

    # ── 出張情報（行3〜6）──
    period = f"{trip_from.strftime('%Y年%m月%d日')}　〜　{trip_to.strftime('%Y年%m月%d日')}"
    info_items = [
        (3, "申請者名", applicant),
        (4, "出張先",   destination),
        (5, "出張目的", purpose),
        (6, "出張期間", period),
    ]
    for r, label, value in info_items:
        ws.merge_cells(f"A{r}:B{r}")
        cl = ws[f"A{r}"]
        cl.value, cl.font, cl.fill, cl.alignment = label, sub_font, sub_fill, l_align
        ws.merge_cells(f"C{r}:G{r}")
        cv = ws[f"C{r}"]
        cv.value, cv.font, cv.alignment = value, n_font, l_align
        ws.row_dimensions[r].height = 20

    # ── 明細ヘッダー（行8）──
    HDR_ROW = 8
    for col, h in enumerate(["No.", "日付", "カテゴリ", "項目名", "金額（円）", "支払方法", "備考"], 1):
        c = ws.cell(row=HDR_ROW, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = h_fill, h_font, c_align, border
    ws.row_dimensions[HDR_ROW].height = 20

    # ── 明細データ（行9〜）──
    for idx, row in enumerate(detail_df.itertuples(index=False), 1):
        r      = HDR_ROW + idx
        amount = int(row[3]) if str(row[3]).replace(".", "").isdigit() else 0
        vals   = [idx, str(row[0]), str(row[1]), str(row[2]),
                  amount, str(row[4]), str(row[5]) if row[5] else ""]
        for col, val in enumerate(vals, 1):
            cell           = ws.cell(row=r, column=col, value=val)
            cell.font      = n_font
            cell.border    = border
            cell.alignment = c_align if col == 1 else (r_align if col == 5 else l_align)
            if col == 5:
                cell.number_format = "#,##0"
        ws.row_dimensions[r].height = 30

    # ── 合計行 ──
    TOT_ROW = HDR_ROW + len(detail_df) + 1
    ws.merge_cells(f"A{TOT_ROW}:D{TOT_ROW}")
    tc = ws[f"A{TOT_ROW}"]
    tc.value, tc.font, tc.alignment, tc.border = "合　計", sub_font, r_align, border_tot_label
    tv = ws.cell(row=TOT_ROW, column=5, value=total_amount)
    tv.font, tv.fill, tv.number_format, tv.alignment, tv.border = sub_font, sub_fill, "#,##0", r_align, border
    for col in [6, 7]:
        ws.cell(row=TOT_ROW, column=col).border = border
    ws.row_dimensions[TOT_ROW].height = 22

    # ── カテゴリ別集計 ──
    SUM_START = TOT_ROW + 2
    ws.merge_cells(f"A{SUM_START}:G{SUM_START}")
    sh = ws[f"A{SUM_START}"]
    sh.value, sh.font, sh.fill, sh.alignment = "カテゴリ別集計", sub_font, sub_fill, l_align
    ws.row_dimensions[SUM_START].height = 18

    SH_ROW = SUM_START + 1
    for col, h in enumerate(["カテゴリ", "合計金額（円）"], 1):
        c = ws.cell(row=SH_ROW, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = h_fill, h_font, c_align, border
    ws.row_dimensions[SH_ROW].height = 15

    for s_idx, row in enumerate(summary_df.itertuples(index=False), 1):
        r = SH_ROW + s_idx
        lc = ws.cell(row=r, column=1, value=row[0])
        lc.font, lc.border, lc.alignment = n_font, border, l_align
        ac = ws.cell(row=r, column=2, value=int(row[1]))
        ac.font, ac.number_format, ac.alignment, ac.border = n_font, "#,##0", r_align, border
        ws.row_dimensions[r].height = 18

    # ── 承認欄 ──
    SIGN_START = SH_ROW + len(summary_df) + 3
    ws.merge_cells(f"A{SIGN_START}:G{SIGN_START}")
    sg = ws[f"A{SIGN_START}"]
    sg.value, sg.font, sg.fill, sg.alignment = "承　認", sub_font, sub_fill, l_align
    ws.row_dimensions[SIGN_START].height = 18

    # 承認者ラベル（E・F・G列に配置）
    for col, label in enumerate(["申請者", "確認者", "承認者"], 5):
        lc = ws.cell(row=SIGN_START + 1, column=col, value=label)
        lc.font, lc.alignment, lc.border = n_font, c_align, border
        sc = ws.cell(row=SIGN_START + 2, column=col)
        sc.border = border
    ws.row_dimensions[SIGN_START + 1].height = 18
    ws.row_dimensions[SIGN_START + 2].height = 48

    # 印刷設定：ページ中央に水平配置
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


excel_buf = build_seisan_excel(
    df, summary, total,
    applicant, destination, purpose, trip_from, trip_to
)
file_name = f"出張経費精算書_{trip_from.strftime('%Y%m%d')}.xlsx"

st.download_button(
    label="📥 精算書Excelをダウンロード",
    data=excel_buf,
    file_name=file_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    type="primary",
)
