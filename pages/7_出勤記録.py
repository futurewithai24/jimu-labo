import streamlit as st
import openpyxl
import pandas as pd
import io
import os
from datetime import datetime, date

st.page_link("home.py", label="← ホームに戻る")
st.title("📬 出勤記録")
st.caption("Outlookから自動収集した出勤メールの記録を確認・ダウンロードできます")

# ============================================================
# ★ 設定：kintai_mail_collector.py と同じパスにしてください
# ============================================================
import pathlib
EXCEL_PATH = str(pathlib.Path.home() / "Desktop" / "出勤記録.xlsx")

# ============================================================
# ファイル読み込み
# ============================================================

def load_workbook_safe(path):
    """Excelファイルを読み込む。存在しない場合はNoneを返す"""
    if not os.path.exists(path):
        return None
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        st.error(f"ファイルの読み込みに失敗しました: {e}")
        return None


def sheet_to_df(ws):
    """シートをDataFrameに変換"""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    df = pd.DataFrame(rows, columns=["氏名", "出勤時刻"])
    df = df.dropna(how="all")
    return df


def sheet_name_to_date(name):
    """'6/10' 形式のシート名をdateに変換（年は当年）"""
    try:
        month, day = map(int, name.split("/"))
        return date(datetime.now().year, month, day)
    except Exception:
        return None


def to_excel_bytes(wb):
    """WorkbookをBytesIOに変換してダウンロード用に返す"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# メイン表示
# ============================================================

wb = load_workbook_safe(EXCEL_PATH)

if wb is None:
    st.info("出勤記録ファイルがまだ作成されていません。\nタスクスケジューラが実行されると自動で作成されます。")
    st.stop()

# シート一覧を日付順に並べる
sheet_names = wb.sheetnames
dated_sheets = []
for name in sheet_names:
    d = sheet_name_to_date(name)
    if d:
        dated_sheets.append((d, name))
dated_sheets.sort(key=lambda x: x[0], reverse=True)  # 新しい日付順

if not dated_sheets:
    st.info("記録されたシートがありません。")
    st.stop()

# ── シート選択 ──────────────────────────────────
st.subheader("📅 日付を選択")

col1, col2 = st.columns([2, 1])
with col1:
    sheet_options = [name for _, name in dated_sheets]
    # デフォルトは今日 or 最新シート
    today_str = f"{datetime.now().month}/{datetime.now().day}"
    default_index = sheet_options.index(today_str) if today_str in sheet_options else 0
    selected_sheet = st.selectbox(
        "表示する日付",
        options=sheet_options,
        index=default_index,
        label_visibility="collapsed"
    )

# ── 選択シートの表示 ────────────────────────────
ws = wb[selected_sheet]
df = sheet_to_df(ws)

st.subheader(f"📋 {selected_sheet} の出勤記録")

if df.empty:
    st.info("この日の記録はありません。")
else:
    st.metric("出勤人数", f"{len(df)} 人")
    st.dataframe(df, use_container_width=True, hide_index=True)

# ── ダウンロードボタン ──────────────────────────
st.divider()
st.subheader("⬇️ ダウンロード")

dl_col1, dl_col2 = st.columns(2)

with dl_col1:
    # 選択日のシートのみDL
    single_wb = openpyxl.Workbook()
    single_ws = single_wb.active
    single_ws.title = selected_sheet

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="Meiryo UI", bold=True, color="FFFFFF")
    for ci, col_name in enumerate(["氏名", "出勤時刻"], 1):
        cell = single_ws.cell(row=1, column=ci, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for ri, row in df.iterrows():
        single_ws.cell(row=ri + 2, column=1, value=row["氏名"])
        single_ws.cell(row=ri + 2, column=2, value=row["出勤時刻"]).alignment = Alignment(horizontal="center")
    single_ws.column_dimensions["A"].width = 20
    single_ws.column_dimensions["B"].width = 12

    st.download_button(
        label=f"📥 {selected_sheet} のみDL",
        data=to_excel_bytes(single_wb),
        file_name=f"出勤記録_{selected_sheet.replace('/', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

with dl_col2:
    # 全シートまとめてDL
    st.download_button(
        label="📥 全日付まとめてDL",
        data=to_excel_bytes(wb),
        file_name="出勤記録_全期間.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# ── シート一覧 ──────────────────────────────────
with st.expander("📂 記録済み日付一覧"):
    for d, name in dated_sheets:
        ws_tmp = wb[name]
        count = ws_tmp.max_row - 1  # ヘッダー除く
        st.write(f"・{name}　（{count}人）")
