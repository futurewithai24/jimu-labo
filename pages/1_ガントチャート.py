import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import io
from datetime import datetime, timedelta
from calendar import monthrange

FONT_NAME = "Meiryo UI"

from utils import show_logo
show_logo()
st.page_link("home.py", label="← ホームに戻る")
st.title("📊 ガントチャート自動生成")
st.caption("CSVをアップロードしてExcel形式のガントチャートを作成します")

# ── サイドバー（設定） ──────────────────────────
with st.sidebar:
    st.header("設定")
    mode = st.selectbox("表示モード", ["daily（3ヶ月以内向け）", "weekly（年間向け）"])
    mode_key = "daily" if mode.startswith("daily") else "weekly"
    bar_color     = st.text_input("ガントバーの色（HEX）", value="4472C4")
    today_color   = st.text_input("今日の列の色（HEX）",   value="FFE699")
    weekend_color = st.text_input("土日の色（HEX・dailyのみ）", value="F2F2F2")

# ── サンプルCSVダウンロード ──────────────────────
sample_csv = """タスク名,開始日,終了日,担当者
企画書作成,2026/06/01,2026/06/10,田中
デザイン確認,2026/06/08,2026/06/20,佐藤
システム開発,2026/06/15,2026/07/10,山田
テスト・修正,2026/07/08,2026/07/20,田中
リリース,2026/07/21,2026/07/25,
"""
st.download_button(
    label="サンプルCSVをダウンロード",
    data=sample_csv.encode("utf-8-sig"),
    file_name="sample_tasks.csv",
    mime="text/csv",
)

st.divider()

# ── CSVアップロード ────────────────────────────
uploaded_file = st.file_uploader("タスクCSVをアップロード（必須列：タスク名・開始日・終了日）", type="csv")

# ── ガントチャート生成ロジック ──────────────────
def parse_date(s):
    for fmt in ('%Y/%m/%d', '%Y-%m-%d'):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(f"日付フォーマットが認識できません: {s}")

def load_tasks_from_upload(file):
    content = file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    tasks = []
    for row in reader:
        tasks.append({
            'task':  row['タスク名'],
            'start': parse_date(row['開始日']),
            'end':   parse_date(row['終了日']),
            'owner': row.get('担当者', ''),
        })
    return tasks

def get_monday(d):
    return d - timedelta(days=d.weekday())

def build_gantt_daily(ws, tasks, today, bar_color, today_color, weekend_color):
    min_date = min(t['start'] for t in tasks).replace(day=1)
    raw_max  = max(t['end']   for t in tasks)
    last_day = monthrange(raw_max.year, raw_max.month)[1]
    max_date = raw_max.replace(day=last_day)
    total_days = (max_date - min_date).days + 1
    COL_TASK  = 1
    COL_OWNER = 2
    COL_DATE0 = 3
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 16
    for col, label in [(COL_TASK, "タスク名"), (COL_OWNER, "担当者")]:
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
    month_start_col = COL_DATE0
    current_month   = None
    for i in range(total_days):
        d   = min_date + timedelta(days=i)
        col = COL_DATE0 + i
        ws.column_dimensions[get_column_letter(col)].width = 2.8
        if d.month != current_month:
            if current_month is not None:
                end_col = col - 1
                if month_start_col < end_col:
                    ws.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=end_col)
                prev_d = min_date + timedelta(days=i - 1)
                cell = ws.cell(row=1, column=month_start_col, value=f"{prev_d.year}年{prev_d.month}月")
                cell.font      = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=9)
                cell.fill      = PatternFill("solid", fgColor="2F5496")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_month   = d.month
            month_start_col = col
    end_col = COL_DATE0 + total_days - 1
    if month_start_col < end_col:
        ws.merge_cells(start_row=1, start_column=month_start_col, end_row=1, end_column=end_col)
    last_d = min_date + timedelta(days=total_days - 1)
    cell = ws.cell(row=1, column=month_start_col, value=f"{last_d.year}年{last_d.month}月")
    cell.font      = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=9)
    cell.fill      = PatternFill("solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16
    for col in [COL_TASK, COL_OWNER]:
        cell = ws.cell(row=2, column=col)
        cell.fill   = PatternFill("solid", fgColor="2F5496")
        cell.border = border
    for i in range(total_days):
        d   = min_date + timedelta(days=i)
        col = COL_DATE0 + i
        cell = ws.cell(row=2, column=col, value=d.day)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        is_weekend = d.weekday() >= 5
        is_today   = (d == today)
        if is_today:
            cell.fill = PatternFill("solid", fgColor=today_color)
            cell.font = Font(name=FONT_NAME, size=8, bold=True, color="333333")
        elif is_weekend:
            cell.fill = PatternFill("solid", fgColor="9DC3E6")
            cell.font = Font(name=FONT_NAME, size=8, color="FFFFFF")
        else:
            cell.fill = PatternFill("solid", fgColor="2F5496")
            cell.font = Font(name=FONT_NAME, size=8, color="FFFFFF")
    for row_idx, task in enumerate(tasks):
        row = 3 + row_idx
        ws.row_dimensions[row].height = 18
        cell = ws.cell(row=row, column=COL_TASK, value=task['task'])
        cell.font      = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(vertical='center')
        cell.border    = border
        cell = ws.cell(row=row, column=COL_OWNER, value=task['owner'])
        cell.font      = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        for i in range(total_days):
            d   = min_date + timedelta(days=i)
            col = COL_DATE0 + i
            cell = ws.cell(row=row, column=col)
            cell.border = border
            in_range   = task['start'] <= d <= task['end']
            is_weekend = d.weekday() >= 5
            is_today   = (d == today)
            if in_range:
                cell.fill = PatternFill("solid", fgColor=bar_color)
            elif is_today:
                cell.fill = PatternFill("solid", fgColor=today_color)
            elif is_weekend:
                cell.fill = PatternFill("solid", fgColor=weekend_color)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 10
    ws.freeze_panes = ws.cell(row=3, column=COL_DATE0)

def build_gantt_weekly(ws, tasks, today, bar_color, today_color):
    min_date = min(t['start'] for t in tasks).replace(day=1)
    raw_max  = max(t['end']   for t in tasks)
    last_day = monthrange(raw_max.year, raw_max.month)[1]
    max_date = raw_max.replace(day=last_day)
    week_start = get_monday(min_date)
    weeks = []
    while week_start <= max_date:
        weeks.append(week_start)
        week_start += timedelta(weeks=1)
    today_week = get_monday(today)
    COL_TASK  = 1
    COL_OWNER = 2
    COL_DATE0 = 3
    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 18
    for col, label in [(COL_TASK, "タスク名"), (COL_OWNER, "担当者")]:
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
    month_groups = {}
    for i, w in enumerate(weeks):
        key = (w.year, w.month)
        month_groups.setdefault(key, []).append(COL_DATE0 + i)
    for (year, month), cols in month_groups.items():
        start_col = min(cols)
        end_col   = max(cols)
        if start_col < end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=f"{year}年{month}月")
        cell.font      = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
    ws.row_dimensions[2].height = 16
    for col in [COL_TASK, COL_OWNER]:
        cell = ws.cell(row=2, column=col)
        cell.fill   = PatternFill("solid", fgColor="2F5496")
        cell.border = border
    for i, w in enumerate(weeks):
        col = COL_DATE0 + i
        ws.column_dimensions[get_column_letter(col)].width = 5.5
        cell = ws.cell(row=2, column=col, value=f"{w.month}/{w.day}")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        if w == today_week:
            cell.fill = PatternFill("solid", fgColor=today_color)
            cell.font = Font(name=FONT_NAME, size=8, bold=True, color="333333")
        else:
            cell.fill = PatternFill("solid", fgColor="2F5496")
            cell.font = Font(name=FONT_NAME, size=8, color="FFFFFF")
    for row_idx, task in enumerate(tasks):
        row = 3 + row_idx
        ws.row_dimensions[row].height = 18
        cell = ws.cell(row=row, column=COL_TASK, value=task['task'])
        cell.font      = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(vertical='center')
        cell.border    = border
        cell = ws.cell(row=row, column=COL_OWNER, value=task['owner'])
        cell.font      = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        for i, w in enumerate(weeks):
            col   = COL_DATE0 + i
            w_end = w + timedelta(days=6)
            cell  = ws.cell(row=row, column=col)
            cell.border = border
            if task['start'] <= w_end and task['end'] >= w:
                cell.fill = PatternFill("solid", fgColor=bar_color)
            elif w == today_week:
                cell.fill = PatternFill("solid", fgColor=today_color)
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 10
    ws.freeze_panes = ws.cell(row=3, column=COL_DATE0)

def generate_excel(tasks, mode_key, bar_color, today_color, weekend_color):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ガントチャート"
    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    if mode_key == "weekly":
        build_gantt_weekly(ws, tasks, today, bar_color, today_color)
    else:
        build_gantt_daily(ws, tasks, today, bar_color, today_color, weekend_color)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ── メイン処理 ──────────────────────────────────
if uploaded_file is not None:
    try:
        tasks = load_tasks_from_upload(uploaded_file)
        st.success(f"{len(tasks)} 件のタスクを読み込みました")
        import pandas as pd
        df = pd.DataFrame([{
            'タスク名': t['task'],
            '開始日': t['start'].strftime('%Y/%m/%d'),
            '終了日': t['end'].strftime('%Y/%m/%d'),
            '担当者': t['owner'],
        } for t in tasks])
        st.dataframe(df, use_container_width=True)
        if st.button("ガントチャートを生成する", type="primary"):
            with st.spinner("作成中です。しばらくお待ちください…"):
                excel_data = generate_excel(tasks, mode_key, bar_color, today_color, weekend_color)
            st.success("完成しました！ダウンロードしてください")
            st.download_button(
                label="Excelをダウンロード",
                data=excel_data,
                file_name="gantt_chart.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as e:
        st.error(f"CSVの列名を確認してください（タスク名・開始日・終了日）")
        st.caption(f"詳細：{e}")

with st.expander("📖 使い方・CSVの書き方"):
    st.write("**CSVに必要な列名（コピーして使ってください）**")
    st.code("タスク名,開始日,終了日,担当者", language=None)
    st.write("- **開始日・終了日** の形式：2026/06/01 または 2026-06-01")
    st.write("- **担当者** は空欄でもOK")
    st.write("- ExcelでCSVを保存するときは「CSV UTF-8（コンマ区切り）」を選択してください")
