import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font
import io

st.page_link("home.py", label="← ホームに戻る")
st.title("📦 在庫チェック自動化")
st.caption("入出庫記録をExcelに反映して、基準値以下の品目を赤くハイライトします")

ALERT_FILL  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
ALERT_FONT  = Font(color="CC0000", bold=True, name="Meiryo UI")
NORMAL_FONT = Font(name="Meiryo UI")

with st.sidebar:
    st.header("設定")
    stock_sheet  = st.text_input("在庫一覧シート名", value="在庫一覧")
    record_sheet = st.text_input("入出庫記録シート名", value="入出庫記録")

uploaded_file = st.file_uploader("在庫管理Excelをアップロード（zaiko_kanri.xlsx）", type=["xlsx"])

if uploaded_file:
    if st.button("在庫を更新する", type="primary"):
        with st.spinner("処理中です…"):
            try:
                wb = openpyxl.load_workbook(uploaded_file)

                if stock_sheet not in wb.sheetnames or record_sheet not in wb.sheetnames:
                    st.error(f"シート名が見つかりません。サイドバーで確認してください。\n存在するシート：{wb.sheetnames}")
                    st.stop()

                ws_stock  = wb[stock_sheet]
                ws_record = wb[record_sheet]

                stock_dict = {}
                for row in ws_stock.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        stock_dict[row[0]] = row[3] if row[3] is not None else 0

                processed_count = 0
                for row in ws_record.iter_rows(min_row=2):
                    flag_cell = row[4]
                    if flag_cell.value == "済":
                        continue
                    item   = row[1].value
                    nyuko  = row[2].value if row[2].value is not None else 0
                    syukko = row[3].value if row[3].value is not None else 0
                    if item and item in stock_dict:
                        stock_dict[item] += nyuko - syukko
                        flag_cell.value = "済"
                        processed_count += 1

                alert_items = []
                for row in ws_stock.iter_rows(min_row=2):
                    name_cell      = row[0]
                    threshold_cell = row[2]
                    stock_cell     = row[3]
                    if not name_cell.value:
                        continue
                    new_stock = stock_dict.get(name_cell.value, 0)
                    stock_cell.value = new_stock
                    stock_cell.font  = NORMAL_FONT
                    threshold = threshold_cell.value if threshold_cell.value is not None else 0
                    if new_stock <= threshold:
                        stock_cell.fill = ALERT_FILL
                        stock_cell.font = ALERT_FONT
                        alert_items.append(f"{name_cell.value}：残り {new_stock}（基準値 {threshold}）")

                out_buf = io.BytesIO()
                wb.save(out_buf)
                out_buf.seek(0)

                st.success(f"完成しました！今回の処理件数：{processed_count} 件")

                if alert_items:
                    st.warning("⚠️ 基準値を下回った品目があります")
                    for item in alert_items:
                        st.write(f"・{item}")
                else:
                    st.info("基準値を下回った品目はありませんでした")

                st.download_button(
                    label="更新済みExcelをダウンロード",
                    data=out_buf,
                    file_name="zaiko_kanri_更新済み.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error(f"エラーが発生しました：{e}")

with st.expander("📖 使い方"):
    st.write("1. 在庫管理Excelをアップロードする（zaiko_kanri.xlsx）")
    st.write("2. 「在庫を更新する」ボタンを押す")
    st.write("3. 更新済みExcelをダウンロードする")
    st.write("- 入出庫記録の未処理行（E列が空）だけを反映します")
    st.write("- 処理済み行には「済」フラグが自動でつきます（二重計算防止）")
    st.write("- 基準値以下の品目は赤くハイライトされます")
    st.write("- シート名が異なる場合はサイドバーで変更してください")
