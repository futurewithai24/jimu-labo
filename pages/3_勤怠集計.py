import streamlit as st
import openpyxl
import pandas as pd
import io
import zipfile
from datetime import timedelta, time, datetime, date

from utils import show_logo
show_logo()
st.page_link("home.py", label="← ホームに戻る")
st.title("📅 勤怠集計")
st.caption("社内の勤怠Excelを複数アップするだけで、社員別の月次集計表を自動作成します（1人1ファイル対応）")

# ── サンプルExcelダウンロード ──────────────────
# 列構成: A=日付, B=曜日, C=開始時刻, D=退勤時刻, E=休憩(h), F=勤務時間数, G=普通残業, H=深夜残業
def _make_kintai_sample():
    from datetime import date, time as t
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    WEEKDAYS = ["月", "火", "水", "木", "金", "土", "日"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "勤怠2026"

    ws.merge_cells("A1:H1")
    ws["A1"] = "勤怠表サンプル（田中 花子）　2026年7月"
    ws["A1"].font = Font(name="Meiryo UI", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["日付", "曜日", "開始時刻", "退勤時刻", "休憩(h)", "勤務時間数", "普通残業", "深夜残業"]
    col_widths = [13, 6, 11, 11, 9, 12, 10, 10]
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hfont = Font(name="Meiryo UI", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    dfont = Font(name="Meiryo UI")
    wkend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    date_fmt = "yyyy/mm/dd"
    time_fmt = "h:mm"

    ot_count = 0
    for day in range(1, 32):
        d = date(2026, 7, day)
        row = 3 + day
        is_weekend = d.weekday() >= 5

        # A: 日付
        dc = ws.cell(row=row, column=1, value=d)
        dc.number_format = date_fmt
        # B: 曜日
        ws.cell(row=row, column=2, value=WEEKDAYS[d.weekday()])

        if not is_weekend:
            ws.cell(row=row, column=3, value=t(9, 0)).number_format = time_fmt   # C: 開始
            ws.cell(row=row, column=4, value=t(18, 0)).number_format = time_fmt  # D: 退勤
            ws.cell(row=row, column=5, value=1.0)                                 # E: 休憩
            ot_count += 1
            if ot_count % 4 == 0:
                ws.cell(row=row, column=6, value=9.0)   # F: 勤務時間数（残業あり）
                ws.cell(row=row, column=7, value=1.0)   # G: 普通残業
            else:
                ws.cell(row=row, column=6, value=8.0)   # F: 勤務時間数

        for ci in range(1, 9):
            cell = ws.cell(row=row, column=ci)
            cell.font = dfont
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if is_weekend:
                cell.fill = wkend_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

st.download_button(
    label="サンプルExcelをダウンロード",
    data=_make_kintai_sample(),
    file_name="勤怠サンプル_田中花子.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
st.caption("※ このサンプルを使って動作確認できます")
st.divider()

# ── ファイルアップロード ────────────────────────
uploaded_files = st.file_uploader(
    "勤怠Excelをアップロード（複数選択OK・1人1ファイル）",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

# ── ファイル解析関数 ────────────────────────────
def td_to_hours(val):
    if val is None:
        return 0.0
    if isinstance(val, timedelta):
        h = val.total_seconds() / 3600
        return round(h, 2) if h > 0 else 0.0
    try:
        h = float(val)
        return round(h, 2) if h > 0 else 0.0
    except Exception:
        return 0.0

KINTAI_SHEET_KEYWORDS = ["オリジナル", "勤怠", "出勤"]

def _get_sheet(wb):
    for name in wb.sheetnames:
        if any(kw in name for kw in KINTAI_SHEET_KEYWORDS):
            return wb[name]
    return wb.active

def parse_kintai(file):
    file.seek(0)
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = _get_sheet(wb)
    sheet_name = ws.title

    name = file.name.rsplit(".", 1)[0]

    attendance = 0
    work_hours = 0.0
    normal_ot  = 0.0
    night_ot   = 0.0

    for row in range(4, 35):
        start = ws.cell(row=row, column=3).value   # C列: 開始時刻
        if start is None:
            continue

        attendance += 1
        work_hours += td_to_hours(ws.cell(row=row, column=6).value)  # F列: 勤務時間数
        normal_ot  += td_to_hours(ws.cell(row=row, column=7).value)  # G列: 普通残業
        night_ot   += td_to_hours(ws.cell(row=row, column=8).value)  # H列: 深夜残業

    return {
        "氏名":          name,
        "出勤日数（日）": attendance,
        "労働時間（h）":  round(work_hours, 1),
        "普通残業（h）":  round(normal_ot, 1),
        "深夜残業（h）":  round(night_ot, 1),
        "_sheet":        sheet_name,
    }

# ── メイン処理 ──────────────────────────────────
if uploaded_files:
    results = []
    errors  = []

    for f in uploaded_files:
        try:
            data = parse_kintai(f)
            results.append(data)
        except Exception as e:
            errors.append(f"{f.name}：{e}")

    if errors:
        for err in errors:
            st.warning(err)

    if results:
        st.success(f"{len(results)} 件のファイルを集計しました")
        df = pd.DataFrame(results)
        with st.expander("🔍 デバッグ情報（問題確認用）"):
            for f2 in uploaded_files:
                f2.seek(0)
                wb2 = openpyxl.load_workbook(f2, data_only=True)
                ws2 = _get_sheet(wb2)
                st.write(f"**ファイル：{f2.name}　シート：{ws2.title}**")
                debug_rows = []
                for r in range(4, 13):
                    c_val = ws2.cell(row=r, column=3).value
                    f_val = ws2.cell(row=r, column=6).value
                    g_val = ws2.cell(row=r, column=7).value
                    debug_rows.append({"行": r, "C列(開始)": c_val, "F列(勤務h)": f_val, "G列(普残)": g_val})
                st.dataframe(pd.DataFrame(debug_rows), use_container_width=True)
        df = df.drop(columns=["_sheet"])
        st.dataframe(df, use_container_width=True)

        def to_excel(df):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "勤怠集計"
            from openpyxl.styles import Font, PatternFill, Alignment
            headers = list(df.columns)
            hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            hfont = Font(name="Meiryo UI", bold=True, color="FFFFFF")
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hfill
                cell.font = hfont
                cell.alignment = Alignment(horizontal="center")
            dfont = Font(name="Meiryo UI")
            for ri, row in enumerate(df.itertuples(index=False), 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val).font = dfont
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        st.download_button(
            label="集計結果をExcelでダウンロード",
            data=to_excel(df),
            file_name="勤怠集計結果.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with st.expander("📖 使い方"):
    st.write("1. 社員ごとの勤怠Excelをまとめてアップロード（複数選択OK）")
    st.write("2. 自動で氏名・出勤日数・労働時間・残業時間を読み取って集計")
    st.write("3. 「集計結果をExcelでダウンロード」ボタンで取得")
    st.write("")
    st.write("**対応フォーマット**")
    st.write("- A列：日付　B列：曜日　C列：開始時刻　D列：退勤時刻　E列：休憩(h)　F列：勤務時間数　G列：普通残業　H列：深夜残業")
    st.write("- 氏名が空の場合はファイル名を氏名として使用します")
