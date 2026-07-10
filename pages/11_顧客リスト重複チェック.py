import streamlit as st
import pandas as pd
import unicodedata
import re
import io
import openpyxl
from openpyxl.styles import PatternFill

st.page_link("home.py", label="← ホームに戻る")
st.title("👥 顧客リスト重複チェック")
st.caption("会社名などのキー列を基に重複行を検出。黄色ハイライト付きExcelでダウンロードできます")

# ── サンプルExcelダウンロード ──────────────────
_sample_customers = pd.DataFrame({
    "会社名":       ["株式会社山田商事", "㈱山田商事", "佐藤物産株式会社", "中村工業", "佐藤物産(株)", "高橋電機"],
    "担当者名":     ["山田 太郎", "山田 太郎", "佐藤 次郎", "中村 三郎", "佐藤 次郎", "高橋 四郎"],
    "メールアドレス": ["yamada@example.com", "yamada@example.com", "sato@example.com",
                     "nakamura@example.com", "sato2@example.com", "takahashi@example.com"],
    "電話番号":     ["03-1234-5678", "03-1234-5678", "03-2345-6789", "03-3456-7890", "03-2345-6789", "03-4567-8901"],
})
_sbuf = io.BytesIO()
_sample_customers.to_excel(_sbuf, sheet_name="顧客リスト", index=False)
_sbuf.seek(0)
st.download_button(
    label="サンプルExcelをダウンロード",
    data=_sbuf,
    file_name="顧客リスト_サンプル.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
st.caption("※ 「株式会社」「㈱」「(株)」のゆれを含む重複サンプルです")
st.divider()


def normalize(text):
    if not isinstance(text, str):
        return text
    text = unicodedata.normalize("NFKC", text)
    text = text.strip()
    text = re.sub(r"㈱|（株）|\(株\)", "株式会社", text)
    text = text.replace("　", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


with st.sidebar:
    st.header("設定")
    auto_delete = st.checkbox("重複行を自動削除する（確認後に使用）", value=False)

uploaded_file = st.file_uploader("Excelファイルをアップロード", type=["xlsx"])

if uploaded_file:
    xl = pd.ExcelFile(uploaded_file)
    sheet_name = st.selectbox("シートを選択", xl.sheet_names)
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, dtype=str)

    st.write(f"**{len(df)} 件のデータを読み込みました**")
    st.dataframe(df.head(), use_container_width=True)

    key_col = st.selectbox("重複チェックのキー列を選択", df.columns.tolist())

    if st.button("重複チェックを実行する", type="primary"):
        with st.spinner("処理中です…"):
            df[f"__norm"] = df[key_col].apply(normalize)
            df["__is_dup"] = df.duplicated(subset=["__norm"], keep="first")
            dup_count = df["__is_dup"].sum()

            if auto_delete:
                df = df[~df["__is_dup"]].copy()

            out_cols = [c for c in df.columns if not c.startswith("__")]
            df_out = df[out_cols].copy()

            buf = io.BytesIO()
            df_out.to_excel(buf, sheet_name=sheet_name, index=False)
            buf.seek(0)

            wb = openpyxl.load_workbook(buf)
            ws = wb[sheet_name]
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            if not auto_delete:
                dup_rows = df[df["__is_dup"]].index.tolist()
                for idx in dup_rows:
                    for cell in ws[idx + 2]:
                        cell.fill = fill

            out_buf = io.BytesIO()
            wb.save(out_buf)
            out_buf.seek(0)

        if auto_delete:
            st.success(f"完成しました！重複 {dup_count} 件を削除しました")
        else:
            st.success(f"完成しました！重複候補 {dup_count} 件を黄色でハイライトしました")

        st.download_button(
            label="Excelをダウンロード",
            data=out_buf,
            file_name="customer_list_checked.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

with st.expander("📖 使い方"):
    st.write("1. Excelファイルをアップロードする")
    st.write("2. シートとキー列（例：会社名）を選択する")
    st.write("3. 「重複チェックを実行する」ボタンを押す")
    st.write("4. 重複行が黄色ハイライトされたExcelをダウンロード")
    st.write("- 全角・半角の違いや（株）表記のゆれも自動で吸収します")
    st.write("- 削除モードは内容を確認してから使ってください")
